import openpyxl
import web3
import datetime

def fetch_data():
    # Set the Ethereum node URL and the contract address
    node_url = "https://mainnet.infura.io/v3/YOUR-PROJECT-ID"
    contract_address = "0x1234567890abcdef1234567890abcdef12345678"

    # Connect to the Ethereum node
    w3 = web3.Web3(web3.Web3.HTTPProvider(node_url))

    # Get the contract instance
    contract = w3.eth.contract(address=contract_address, abi=CONTRACT_ABI)

    # Define the start and end dates for the data range (3 months ago to now)
    start_date = datetime.datetime.now() - datetime.timedelta(days=90)
    end_date = datetime.datetime.now()

    # Fetch the data from the contract
    data = []
    for i in range(start_date, end_date):
        data.append(contract.functions.getData(i).call())

    return data

def publish_data(data):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Add a new worksheet
    worksheet = workbook.create_sheet("Data")

    # Write the data to the worksheet, starting at cell A1
    for i, row in enumerate(data):
        for j, cell in enumerate(row):
            worksheet.cell(row=i+1, column=j+1, value=cell)

    # Save the workbook
    workbook.save("data.xlsx")

if __name__ == "__main__":
    # Fetch the data
    data = fetch_data()

    # Publish the data
    publish_data(data)
